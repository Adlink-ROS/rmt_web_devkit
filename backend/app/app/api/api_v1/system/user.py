from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi import File, UploadFile
from sqlalchemy.orm import Session, joinedload
# excel
from io import BytesIO
from openpyxl import load_workbook

from app import models, schemas
from app.api import deps
from app.core.config import settings
from app.core.security import get_password_hash
from app.extensions.utils import list_to_tree

router = APIRouter()


# @router.get("/me", response_model=schemas.Response)
# def read_user_me(current_user: models.User = Depends(deps.get_current_active_user)) -> Any:
#     user = current_user.dict()
#     user['roles'] = [role.role.id for role in current_user.user_role]
#     return {"code": 20000, "data": user}
@router.get("/me", response_model=schemas.Response)
def read_user_me() -> Any:
    user = {'avatar': 'https://wpimg.wallstcn.com/f778738c-e4f8-4870-b634-56703b4acafe.gif'}
    user['roles'] = [1]
    return {"code": 20000, "data": user}


# @router.get("/list", response_model=schemas.Response)
# def read_users(*,
#                db: Session = Depends(deps.get_db),
#                limit: int,
#                page: int,
#                deptId: Optional[int] = None,
#                username: Optional[str] = None,
#                nickname: Optional[str] = None,
#                status: Optional[str] = None, ) -> Any:
#     """用户管理-查询"""
#     query = db.query(models.User, models.Department).outerjoin(
#         models.User_Department, models.User_Department.user_id == models.User.id).outerjoin(
#         models.Department, models.Department.id == models.User_Department.department_id)
#     if username: query = query.filter(models.User.username.like("%" + username + "%"))
#     if nickname: query = query.filter(models.User.nickname.like("%" + nickname + "%"))
#     if status: query = query.filter(models.User.status == status)
#     # 根据部门ID筛选部门及部门下级所有员工
#     if deptId:
#         departments = db.query(models.Department).filter(models.Department.id, models.Department.status == 1).all()
#         tree = list_to_tree([dep.dict() for dep in departments], root_id=deptId)

#         # 递归获取部门及部门下级
#         def get_list_id_by_tree(nodes):
#             ids = [nodes["id"], ]
#             if nodes.get("children"):
#                 for node in nodes["children"]: ids = ids + get_list_id_by_tree(node)
#             return ids

#         tree_ids = get_list_id_by_tree(tree)
#         query = query.filter(models.User_Department.department_id.in_(tree_ids))

#     total = query.count()
#     users = query.order_by(models.User.username).limit(limit).offset((page - 1) * limit).all()

#     user_list = []
#     for user in users:
#         user_info = user[0].dict()
#         user_info["department"] = user[1].dict()
#         user_list.append(user_info)
#     return {"code": 20000, "data": {"items": user_list, 'total': total}, }


# @router.get("/", response_model=schemas.Response)
# def read_user(*, db: Session = Depends(deps.get_db)) -> Any:
#     """用户管理-新增前获取role和post"""
#     roleOptions = db.query(models.Role).all()
#     postOptions = db.query(models.Dict_Data).join(
#         models.Dict_Type, models.Dict_Type.id == models.Dict_Data.type_id).filter(
#         models.Dict_Type.code == "post").all()
#     return {"code": 20000, "data": {"roleOptions": roleOptions, "postOptions": postOptions}}


# @router.get("/{id}", response_model=schemas.Response)
# def read_user(*, db: Session = Depends(deps.get_db), id: int) -> Any:
#     """用户管理-修改前根据id查询"""
#     # 角色
#     roleOptions = db.query(models.Role).all()
#     # 岗位
#     postOptions = db.query(models.Dict_Data).join(
#         models.Dict_Type, models.Dict_Type.id == models.Dict_Data.type_id).filter(
#         models.Dict_Type.code == "post").all()
#     # 用户
#     user = db.query(models.User).filter(models.User.id == id).options(
#         joinedload(models.User.user_department), joinedload(models.User.user_role)).one()
#     user_post = db.query(models.User_Dict).outerjoin(
#         models.Dict_Data, models.Dict_Data.id == models.User_Dict.dict_id).outerjoin(
#         models.Dict_Type, models.Dict_Type.id == models.Dict_Data.type_id).filter(
#         models.Dict_Type.code == "post", models.User_Dict.user_id == id).all()

#     user_info = user.dict()
#     user_info["deptId"] = user.user_department[0].department_id
#     user_info["roleIds"] = [user_role.role_id for user_role in user.user_role]
#     user_info["postIds"] = [post.dict_id for post in user_post]
#     return {"code": 20000, "data": {"user": user_info, "roleOptions": roleOptions, "postOptions": postOptions}}


# @router.put("/", response_model=schemas.Response)
# def update_user(*, db: Session = Depends(deps.get_db), user: schemas.UserUpdate) -> Any:
#     """用户管理-修改"""
#     user_id = user.id

#     user = user.dict()
#     deptId = user.pop("deptId")
#     postIds = user.pop("postIds")
#     roleIds = user.pop("roleIds")
#     # User
#     db.query(models.User).filter(models.User.id == user_id).update(user)
#     db.flush()
#     # department
#     db.query(models.User_Department).filter(models.User_Department.user_id == user_id).delete()
#     user_department = {"user_id": user_id, "department_id": deptId}
#     db.add(models.User_Department(**user_department))
#     db.flush()
#     # role
#     db.query(models.User_Role).filter(models.User_Role.user_id == user_id).delete()
#     user_roles = [{"user_id": user_id, "role_id": i} for i in roleIds]
#     db.bulk_insert_mappings(models.User_Role, user_roles)
#     db.flush()
#     # dict
#     db.query(models.User_Dict).filter(models.User_Dict.user_id == user_id).delete()
#     # post
#     user_post = [{"user_id": user_id, "dict_id": i} for i in postIds]
#     user_dict = user_post + []
#     db.bulk_insert_mappings(models.User_Dict, user_dict)
#     db.flush()
#     return {"code": 20000, "message": "修改成功"}


# @router.post("/", response_model=schemas.Response)
# def add_user(*, db: Session = Depends(deps.get_db), user: schemas.UserCreate) -> Any:
#     """用户管理-新增"""
#     user = user.dict()
#     deptId = user.pop("deptId")
#     postIds = user.pop("postIds")
#     roleIds = user.pop("roleIds")
#     # User
#     add_user = models.User(**user)
#     db.add(add_user)
#     db.flush()
#     # department
#     user_department = {"user_id": add_user.id, "department_id": deptId, }
#     db.add(models.User_Department(**user_department))
#     db.flush()
#     # role
#     user_roles = [{"user_id": add_user.id, "role_id": i} for i in roleIds]
#     db.bulk_insert_mappings(models.User_Role, user_roles)
#     db.flush()
#     # dict
#     # post
#     user_post = [{"user_id": add_user.id, "dict_id": i} for i in postIds]
#     user_dict = user_post + []
#     db.bulk_insert_mappings(models.User_Dict, user_dict)
#     db.flush()
#     return {"code": 20000, "message": "新增成功", }


# @router.put("/reset-password", response_model=schemas.Response)
# def reset_password(*,
#                    User=Depends(deps.get_current_active_user),
#                    db: Session = Depends(deps.get_db),
#                    reset: schemas.UserPWReset
#                    ) -> Any:
#     data = {"hashed_password": get_password_hash(reset.password)}
#     # Only superuser can reset the password
#     if User.is_superuser or User.id == reset.user_id:
#         db.query(models.User).filter(models.User.id == reset.user_id).update(data)
#         return {"code": 20000, "message": "success"}
#     raise HTTPException(status_code=400, detail="permission denied")


# @router.delete("/{ids}", response_model=schemas.Response)
# def delete_user(*, db: Session = Depends(deps.get_db), ids: str) -> Any:
#     """用户管理-删除用户"""
#     ids = [int(id) for id in ids.split(",")]
#     db.query(models.User).filter(models.User.id.in_(ids)).delete(synchronize_session=False)
#     return {"code": 20000, "message": "删除成功", }


# @router.post("/importData", response_model=schemas.Response)
# def create_file(db: Session = Depends(deps.get_db), updateSupport: bool = False, file: UploadFile = File(...)):
#     def check_dict_label(label, code):
#         dict = db.query(models.Dict_Data).outerjoin(
#             models.Dict_Type, models.Dict_Type.id == models.Dict_Data.type_id).filter(
#             models.Dict_Data.label == label, models.Dict_Type.code == code).one()
#         return dict

#     try:
#         io = BytesIO(file.file.read())
#         wb = load_workbook(io, read_only=True)
#         ws = wb.active  # wb.worksheets[0]
#         for row in ws.iter_rows(min_row=2):
#             # dict_data
#             sex = check_dict_label(row[5].value.strip(""), "sex").label
#             status = check_dict_label(row[6].value.strip(""), "user_status").label

#             user = {
#                 "username": row[0].value.strip(""),
#                 "nickname": row[1].value.strip(""),
#                 "identity_card": row[3].value.strip(""),
#                 "phone": row[4].value.strip(""),
#                 "sex": sex,
#                 "status": status,
#                 "hashed_password": get_password_hash(settings.INIT_PASSWORD)
#             }
#             department = db.query(models.Department).filter(models.Department.name == row[2].value.strip("")).one()
#             posts = db.query(models.Dict_Data).outerjoin(
#                 models.Dict_Type, models.Dict_Type.id == models.Dict_Data.type_id).filter(
#                 models.Dict_Data.label.in_(row[7].value.strip("").split(",")), models.Dict_Type.code == "post").all()
#             exist_user = db.query(models.User).filter(models.User.username == user["username"])

#             if not exist_user.first():
#                 user = models.User(**user)
#                 db.add(user)
#                 db.flush()
#                 user_department = {"user_id": user.id, "department_id": department.id}
#                 db.add(models.User_Department(**user_department))
#                 user_dict = [{"user_id": user.id, "dict_id": post.id} for post in posts]
#                 db.bulk_insert_mappings(models.User_Dict, user_dict)
#             elif updateSupport:
#                 exist_user_id = exist_user.one().id
#                 exist_user.update(user)
#                 db.flush()
#                 # department
#                 db.query(models.User_Department).filter(models.User_Department.user_id == exist_user_id).delete()
#                 user_department = {"user_id": exist_user_id, "department_id": department.id}
#                 db.add(models.User_Department(**user_department))
#                 # post
#                 db.query(models.User_Dict).filter(models.User_Dict.user_id == exist_user_id).delete()
#                 user_dict = [{"user_id": exist_user_id, "dict_id": post.id} for post in posts]
#                 db.bulk_insert_mappings(models.User_Dict, user_dict)
#         return {"code": 20000, "message": "导入成功"}
#     except Exception as exc:
#         raise HTTPException(status_code=200, detail=f"导入失败，请检查数据!   Error Reason: {exc}")
#     finally:
#         wb.close()
